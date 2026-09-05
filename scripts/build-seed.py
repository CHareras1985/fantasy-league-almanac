"""
Builds data/seed/*.json (2009-2021 reconciled history) from an .xlsx export of
the league's Google Sheet. See data/seed/README.md for the output schema and
fantasy-league-project-brief.md (Simple Picks repo) for the reconciliation this
encodes (name normalization, the 4 corrected award-point tallies, etc).

Usage:
    python3 scripts/build-seed.py path/to/league-export.xlsx

Get the .xlsx export with:
    curl -sL -o league-export.xlsx \\
      "https://docs.google.com/spreadsheets/d/1SGatoxS0HlpQf-VbVzTwc8dg6eHMeHtndZmbhDwA_5Y/export?format=xlsx"
"""
import openpyxl, json, re, os, sys

if len(sys.argv) != 2:
    print(__doc__)
    sys.exit(1)

SRC = sys.argv[1]
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "seed")
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------- naming ----
MGR_ALIAS = {
    'fern': 'fernando', 'fernando': 'fernando',
    'marcil': 'marcil', 'george': 'george', 'dean': 'dean', 'tony': 'tony',
    'costa': 'costa', 'fink': 'fink', 'chris': 'chris', 'matt': 'matt',
    'mike': 'mike', 'wally': 'wally', 'paul': 'paul',
}
MGR_DISPLAY = {
    'george': 'George', 'dean': 'Dean', 'fernando': 'Fernando', 'tony': 'Tony',
    'marcil': 'Marcil', 'costa': 'Costa', 'fink': 'Fink', 'chris': 'Chris',
    'matt': 'Matt', 'mike': 'Mike', 'wally': 'Wally', 'paul': 'Paul',
}
PLACEHOLDER = {'random', 'rando', 'zero made', ''}

def mgr(n):
    if n is None:
        return None
    s = str(n).strip()
    if s.lower() in PLACEHOLDER:
        return None
    return MGR_ALIAS.get(s.lower(), s.lower())

PLAYER_ALIAS = {
    'aaron roders': 'Aaron Rodgers',
    'david alers': 'David Akers',
    'marr forte': 'Matt Forte',
    'kirk cousins': 'Kirk Cousins',
    'kirk cousins ': 'Kirk Cousins',
    'kirk cousins': 'Kirk Cousins',
    'mathew stafford': 'Matthew Stafford',
    'alson jeffery': 'Alshon Jeffery',
    'sabastian janikowski': 'Sebastian Janikowski',
    'saquan barkley': 'Saquon Barkley',
    'devin funches': 'Devin Funchess',
    'montey ball': 'Montee Ball',
    'dionte johnson': 'Diontae Johnson',
    'devante adams': 'Davante Adams',
    'devante adams': 'Davante Adams',
    'piere garcon': 'Pierre Garcon',
    "jones-drew": "Maurice Jones-Drew",
}
# Yahoo/dump typos with unusual casing need exact-match keys; add case-insensitive lookup
def norm_player(p):
    if p is None:
        return None
    s = str(p).strip()
    if s == '' or s.lower() in PLACEHOLDER:
        return None
    key = s.lower().rstrip()
    if key in PLAYER_ALIAS:
        return PLAYER_ALIAS[key]
    if key == 'kirk cousins'.lower():
        return 'Kirk Cousins'
    if key == 'kirk cousins':
        return 'Kirk Cousins'
    # trailing-space / stray-space typos e.g. "Cam Newton "
    return re.sub(r'\s+', ' ', s).strip()

def num(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace(',', '').strip()
        if v == '':
            return None
        try:
            return float(v)
        except ValueError:
            return v
    return float(v)

def cell(ws, addr):
    return ws[addr].value

SEASON_TABS = [s for s in wb.sheetnames if s.startswith('Season ')]
def year_of(tab):
    return 2009 + int(re.search(r'\((\d+)\)', tab).group(1)) - 1
SEASON_TABS.sort(key=year_of)

# --------------------------------------------------------- award corrections ----
# (season_index_key -> manager -> corrected total) from the reconciliation pass
CORRECTIONS = {
    2010: {'marcil': 12},
    2012: {'marcil': 13},
    2018: {'mike': 8},
    2021: {'matt': 1},
}

CATS = {
    'reg_season_1st': 4, 'reg_season_2nd': 3, 'reg_season_3rd': 2,
    'winning_season': 1,
    'playoff_champion': 5, 'playoff_runner_up': 4, 'playoff_third': 3,
    'most_points': 3, 'mvp_pick': 2, 'best_transaction': 2, 'longest_streak': 3,
}

managers_seen = {}     # slug -> {first, last, seasons}
seasons_out = []
standings_out = []
playoffs_out = []
season_awards_out = []
draft_highlights_out = []
all_star_out = []
transactions_out = []

for tab in SEASON_TABS:
    ws = wb[tab]
    yr = year_of(tab)
    late = yr >= 2018

    # ---- standings ----
    rows = []
    for r in range(2, 12):
        nm = mgr(cell(ws, f'A{r}'))
        w, l, pts = cell(ws, f'B{r}'), cell(ws, f'C{r}'), cell(ws, f'D{r}')
        if nm and w is not None:
            rows.append({'manager': nm, 'w': num(w), 'l': num(l), 'points_for': num(pts)})
    ranked = sorted(rows, key=lambda x: (-x['w'], -(x['points_for'] or 0)))
    for i, row in enumerate(ranked):
        row['reg_rank'] = i + 1

    team_count = len(rows)
    notes = None
    if yr in (2009, 2010):
        notes = f"{team_count}-team season; one team played a phantom 'Random' (median) opponent some weeks, so league W-L does not balance"
    regular_weeks = int(round(sum(r['w'] + r['l'] for r in rows) / team_count)) if team_count else None

    seasons_out.append({
        'year': yr, 'team_count': team_count, 'regular_weeks': regular_weeks, 'notes': notes,
    })

    for row in ranked:
        m = row['manager']
        managers_seen.setdefault(m, {'first': yr, 'last': yr})
        managers_seen[m]['first'] = min(managers_seen[m]['first'], yr)
        managers_seen[m]['last'] = max(managers_seen[m]['last'], yr)
        standings_out.append({
            'season': yr, 'manager': m, 'wins': row['w'], 'losses': row['l'],
            'points_for': row['points_for'], 'reg_rank': row['reg_rank'],
        })

    # ---- playoffs ----
    champ, runner, third = mgr(cell(ws, 'G2')), mgr(cell(ws, 'G3')), mgr(cell(ws, 'G4'))
    champ_a, champ_a_pts = mgr(cell(ws, 'F7')), num(cell(ws, 'G7'))
    champ_b, champ_b_pts = mgr(cell(ws, 'F8')), num(cell(ws, 'G8'))
    third_a, third_a_pts = mgr(cell(ws, 'F11')), num(cell(ws, 'G11'))
    third_b, third_b_pts = mgr(cell(ws, 'F12')), num(cell(ws, 'G12'))
    playoffs_out.append({
        'season': yr,
        'champion': champ, 'runner_up': runner, 'third': third,
        'championship_game': (
            [{'manager': champ_a, 'points': champ_a_pts}, {'manager': champ_b, 'points': champ_b_pts}]
            if champ_a_pts is not None else None
        ),
        'third_place_game': (
            [{'manager': third_a, 'points': third_a_pts}, {'manager': third_b, 'points': third_b_pts}]
            if third_a_pts is not None else None
        ),
    })

    # ---- best / worst draft pick, best transaction (per manager) ----
    for r in range(2, 12):
        name = mgr(cell(ws, f'A{r}'))
        if not name:
            continue
        bp_player, bp_pts = norm_player(cell(ws, f'K{r}')), num(cell(ws, f'L{r}'))
        if bp_player:
            draft_highlights_out.append({
                'season': yr, 'manager': name, 'kind': 'best',
                'player': bp_player, 'player_points': bp_pts,
            })
        if not late:
            wp_player, wp_pts = norm_player(cell(ws, f'O{r}')), num(cell(ws, f'P{r}'))
            if wp_player:
                draft_highlights_out.append({
                    'season': yr, 'manager': name, 'kind': 'worst',
                    'player': wp_player, 'player_points': wp_pts,
                })
            tx_player, tx_pts = norm_player(cell(ws, f'S{r}')), num(cell(ws, f'T{r}'))
        else:
            tx_player, tx_pts = norm_player(cell(ws, f'O{r}')), num(cell(ws, f'P{r}'))
        if tx_player:
            transactions_out.append({
                'season': yr, 'manager': name, 'player': tx_player, 'player_points': tx_pts,
            })

    # ---- all-star roster ----
    if not late:
        pos_col, player_col, pts_col, owner_col = 'V', 'W', 'X', 'Y'
    else:
        pos_col, player_col, pts_col, owner_col = 'R', 'S', 'T', 'U'
    for r in range(2, 12):
        pos = cell(ws, f'{pos_col}{r}')
        player = norm_player(cell(ws, f'{player_col}{r}'))
        pts = num(cell(ws, f'{pts_col}{r}'))
        owner = mgr(cell(ws, f'{owner_col}{r}'))
        if pos and player:
            all_star_out.append({
                'season': yr, 'position': str(pos).strip(), 'player': player,
                'player_points': pts, 'manager': owner,
            })

    # ---- season-level awards ----
    mvp_mgr = mgr(cell(ws, 'K13'))
    mvp_player, mvp_pts = norm_player(cell(ws, 'K14')), num(cell(ws, 'L14'))
    mvp_round, mvp_pick = num(cell(ws, 'K15')), num(cell(ws, 'K16'))

    if not late:
        trans_mgr = mgr(cell(ws, 'S13'))
        trans_player, trans_pts = norm_player(cell(ws, 'S14')), num(cell(ws, 'T14'))
    else:
        trans_mgr = mgr(cell(ws, 'O13'))
        trans_player, trans_pts = norm_player(cell(ws, 'O14')), num(cell(ws, 'P14'))

    most_pts_mgr, most_pts_val = mgr(cell(ws, 'G14')), num(cell(ws, 'G15'))
    fewest_pts_mgr, fewest_pts_val = mgr(cell(ws, 'I14')), num(cell(ws, 'I15'))
    streak_mgrs = [x for x in [mgr(cell(ws, 'K18')), mgr(cell(ws, 'L18'))] if x]
    streak_len = num(cell(ws, 'K19'))

    reg1 = ranked[0]['manager'] if len(ranked) > 0 else None
    reg2 = ranked[1]['manager'] if len(ranked) > 1 else None
    reg3 = ranked[2]['manager'] if len(ranked) > 2 else None

    # compute award points per manager (source of truth going forward)
    pts_by_mgr = {row['manager']: 0 for row in ranked}
    def add(m, key):
        if not m:
            return
        pts_by_mgr[m] = pts_by_mgr.get(m, 0) + CATS[key]
    if reg1: add(reg1, 'reg_season_1st')
    if reg2: add(reg2, 'reg_season_2nd')
    if reg3: add(reg3, 'reg_season_3rd')
    for row in ranked:
        if row['w'] is not None and row['l'] is not None and row['w'] > row['l']:
            add(row['manager'], 'winning_season')
    if champ: add(champ, 'playoff_champion')
    if runner: add(runner, 'playoff_runner_up')
    if third: add(third, 'playoff_third')
    if most_pts_mgr: add(most_pts_mgr, 'most_points')
    if mvp_mgr: add(mvp_mgr, 'mvp_pick')
    if trans_mgr: add(trans_mgr, 'best_transaction')
    for sm in streak_mgrs:
        add(sm, 'longest_streak')

    # apply confirmed corrections (override tally slips found during reconciliation)
    for m, corrected in CORRECTIONS.get(yr, {}).items():
        pts_by_mgr[m] = corrected

    for m, total in pts_by_mgr.items():
        season_awards_out.append({
            'season': yr, 'manager': m, 'points': total, 'source': 'imported',
            'detail': {
                'reg_rank': next((row['reg_rank'] for row in ranked if row['manager'] == m), None),
                'winning_season': any(row['manager'] == m and row['w'] > row['l'] for row in ranked),
                'playoff_finish': (
                    'champion' if m == champ else 'runner_up' if m == runner else 'third' if m == third else None
                ),
                'most_points': m == most_pts_mgr,
                'mvp_pick': m == mvp_mgr,
                'best_transaction': m == trans_mgr,
                'longest_streak': m in streak_mgrs,
                'streak_length': streak_len if m in streak_mgrs else None,
            },
        })

    if mvp_mgr:
        draft_highlights_out.append({
            'season': yr, 'manager': mvp_mgr, 'kind': 'mvp',
            'player': mvp_player, 'player_points': mvp_pts,
            'draft_round': mvp_round, 'draft_pick': mvp_pick,
        })

print(f"Seasons: {len(seasons_out)}")
print(f"Standings rows: {len(standings_out)}")
print(f"Playoff rows: {len(playoffs_out)}")
print(f"Season award rows: {len(season_awards_out)}")
print(f"Draft highlight rows: {len(draft_highlights_out)}")
print(f"All-star rows: {len(all_star_out)}")
print(f"Transaction rows: {len(transactions_out)}")
print(f"Managers: {sorted(managers_seen)}")

# ---------------------------------------------------------------- managers ----
managers_out = []
for slug, span in sorted(managers_seen.items(), key=lambda x: x[1]['first']):
    managers_out.append({
        'id': slug,
        'name': MGR_DISPLAY[slug],
        'aliases': ['Fern'] if slug == 'fernando' else [],
        'first_season': span['first'],
        'last_season': span['last'],
        'active': span['last'] == 2021,   # recomputed once 2022-2025 land from Yahoo
    })

award_rules_out = [
    {'key': 'playoff_champion', 'points': 5, 'label': 'Playoff Champion'},
    {'key': 'playoff_runner_up', 'points': 4, 'label': 'Playoff Runner-up'},
    {'key': 'reg_season_1st', 'points': 4, 'label': 'Regular Season 1st'},
    {'key': 'playoff_third', 'points': 3, 'label': 'Playoff 3rd Place'},
    {'key': 'reg_season_2nd', 'points': 3, 'label': 'Regular Season 2nd'},
    {'key': 'most_points', 'points': 3, 'label': 'Most PTS Scored'},
    {'key': 'longest_streak', 'points': 3, 'label': 'Longest Win Streak'},
    {'key': 'reg_season_3rd', 'points': 2, 'label': 'Regular Season 3rd'},
    {'key': 'mvp_pick', 'points': 2, 'label': 'M.V.P. Pick'},
    {'key': 'best_transaction', 'points': 2, 'label': 'Best Transaction'},
    {'key': 'winning_season', 'points': 1, 'label': 'Winning Season'},
]

files = {
    'managers.json': managers_out,
    'seasons.json': seasons_out,
    'standings.json': standings_out,
    'playoff_results.json': playoffs_out,
    'award_rules.json': award_rules_out,
    'season_awards.json': season_awards_out,
    'draft_highlights.json': draft_highlights_out,
    'all_star_slots.json': all_star_out,
    'transactions.json': transactions_out,
}
for name, data in files.items():
    with open(os.path.join(OUT, name), 'w') as f:
        json.dump(data, f, indent=2, default=str)
    print(f"wrote {name} ({len(data)} records)" if isinstance(data, list) else f"wrote {name}")
